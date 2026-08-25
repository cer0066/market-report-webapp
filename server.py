from __future__ import annotations

import copy
import json
import os
import re
import shutil
import sys
import threading
import traceback
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time
from email.parser import BytesParser
from email.policy import default as email_policy
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote, urlparse

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR = BASE_DIR / "uploads"
SAMPLE_DIR = BASE_DIR.parent
JOBS: dict[str, dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()

SHEET_UNITS = {
    "水电": ("俄日", "红卫桥"),
    "光伏": ("腾龙", "凤舞", "赛日"),
}

PLANT_ALIASES = {
    "俄日": "俄日",
    "红卫桥": "红卫桥",
    "腾龙": "腾龙",
    "腾龙光伏电站": "腾龙",
    "凤舞": "凤舞",
    "凤舞光伏电站": "凤舞",
    "赛日": "赛日",
    "赛日光伏电站": "赛日",
}

MARKET_LABELS = {
    "priority": "省内优先",
    "retention": "留存",
    "province": "省内市场",
    "export": "省间外送",
}

PV_RESET_MARKETS = tuple(MARKET_LABELS.values())
NEAR_ZERO_QUANTITY = 1e-6
SUMMARY_LABELS = {"汇总", "合计", "总计", "总总"}

INPUT_UNITS = {
    "hydro_contract": ("俄日", "红卫桥"),
    "hydro_priority": ("俄日", "红卫桥"),
    "hydro_retention": ("俄日", "红卫桥"),
    "hydro_export": ("俄日", "红卫桥"),
    "pv_priority_tenglong_fengwu": ("腾龙", "凤舞"),
    "pv_retention_tenglong_fengwu": ("腾龙", "凤舞"),
    "pv_contract_tenglong_fengwu": ("腾龙", "凤舞"),
    "pv_export_tenglong_fengwu": ("腾龙", "凤舞"),
    "pv_priority_sairi": ("赛日",),
    "pv_retention_sairi": ("赛日",),
    "pv_contract_sairi": ("赛日",),
    "pv_export_sairi": ("赛日",),
}

PROTECTED_RANGES_BY_SHEET = {
    "水电": ("L:P", "AC:AG"),
    "光伏": ("L:P", "AC:AG", "AT:AX"),
}

REQUIRED_FILES = {
    "template": "最终报送表模板",
}

SAMPLE_FILES = {
    "template": SAMPLE_DIR / "8月13日现货市场出清情况-金川.xlsx",
    "hydro_contract": SAMPLE_DIR / "中长期合同明细 (4)(1).xlsx",
    "hydro_priority": SAMPLE_DIR / "转让后优先电量 (3).xlsx",
    "hydro_retention": SAMPLE_DIR / "转让后留存电量 (1).xlsx",
    "hydro_export": SAMPLE_DIR / "转让后外送 (3).xlsx",
}

SAMPLE_87_DIR = SAMPLE_DIR / "8.7"
SAMPLE_87_FILES = {
    "template": SAMPLE_87_DIR / "（最终报送表）8月7日现货市场出清情况-金川.xlsx",
    "pv_contract_sairi": SAMPLE_87_DIR / "赛日8月7日（平台导出表，无省间现货）" / "中长期合同明细.xlsx",
    "pv_contract_tenglong_fengwu": SAMPLE_87_DIR / "腾龙、凤舞8月7日（平台导出表，无省间现货）" / "中长期合同明细.xlsx",
    "pv_export_tenglong_fengwu": SAMPLE_87_DIR / "腾龙、凤舞8月7日（平台导出表，无省间现货）" / "转让后外送.xlsx",
}


@dataclass(frozen=True)
class SourceSpec:
    key: str
    title: str
    market_key: str
    unit_col: str
    date_col: str
    slot_col: str
    quantity_col: str
    price_col: str
    green_col: str | None = None
    preferred_sheet: str | None = None


SOURCE_SPECS = {
    "contract": SourceSpec(
        key="contract",
        title="中长期合同明细",
        market_key="province",
        unit_col="售方主体",
        date_col="交易标的日期",
        slot_col="交易时段",
        quantity_col="合同电量",
        price_col="合同电价",
        green_col="绿证价格",
        preferred_sheet="中长期合同明细",
    ),
    "priority": SourceSpec(
        key="priority",
        title="转让后优先电量",
        market_key="priority",
        unit_col="交易单元",
        date_col="交易标的日期",
        slot_col="交易时段",
        quantity_col="优先计划电量",
        price_col="优先计划电价",
        preferred_sheet="优先电量",
    ),
    "retention": SourceSpec(
        key="retention",
        title="转让后留存电量",
        market_key="retention",
        unit_col="交易单元",
        date_col="交易标的日期",
        slot_col="交易时段",
        quantity_col="留存电量",
        price_col="留存电价",
        preferred_sheet="留存电量",
    ),
    "export": SourceSpec(
        key="export",
        title="转让后外送",
        market_key="export",
        unit_col="交易单元",
        date_col="交易标的日期",
        slot_col="交易时段",
        quantity_col="外送电量",
        price_col="外送电价",
        green_col="绿证价格",
        preferred_sheet="外送电量",
    ),
}

SOURCE_INPUTS = {
    "hydro_contract": {"sheet": "水电", "spec": "contract"},
    "hydro_priority": {"sheet": "水电", "spec": "priority"},
    "hydro_retention": {"sheet": "水电", "spec": "retention"},
    "hydro_export": {"sheet": "水电", "spec": "export"},
    "pv_priority_tenglong_fengwu": {"sheet": "光伏", "spec": "priority"},
    "pv_retention_tenglong_fengwu": {"sheet": "光伏", "spec": "retention"},
    "pv_contract_sairi": {"sheet": "光伏", "spec": "contract"},
    "pv_contract_tenglong_fengwu": {"sheet": "光伏", "spec": "contract"},
    "pv_export_tenglong_fengwu": {"sheet": "光伏", "spec": "export"},
    "pv_priority_sairi": {"sheet": "光伏", "spec": "priority"},
    "pv_retention_sairi": {"sheet": "光伏", "spec": "retention"},
    "pv_export_sairi": {"sheet": "光伏", "spec": "export"},
}

OPTIONAL_INPUT_LABELS = {
    "hydro_priority": "水电 俄日/红卫桥 省内优先",
    "hydro_retention": "水电 俄日/红卫桥 留存",
    "hydro_contract": "水电 俄日/红卫桥 省内市场",
    "hydro_export": "水电 俄日/红卫桥 省间外送",
    "pv_priority_tenglong_fengwu": "光伏 腾龙/凤舞 省内优先",
    "pv_retention_tenglong_fengwu": "光伏 腾龙/凤舞 留存",
    "pv_contract_tenglong_fengwu": "光伏 腾龙/凤舞 省内市场",
    "pv_export_tenglong_fengwu": "光伏 腾龙/凤舞 省间外送",
    "pv_priority_sairi": "光伏 赛日 省内优先",
    "pv_retention_sairi": "光伏 赛日 留存",
    "pv_contract_sairi": "光伏 赛日 省内市场",
    "pv_export_sairi": "光伏 赛日 省间外送",
}


class AppError(Exception):
    pass


def to_number(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text == "" or text in {"-", "--"}:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace("/", "-").replace(".", "-")
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if not match:
        return text
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def normalize_time(value: Any) -> str:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    text = str(value or "").strip()
    match = re.search(r"(\d{1,2}):(\d{2})", text)
    if not match:
        return text
    hour, minute = match.groups()
    return f"{int(hour):02d}:{int(minute):02d}"


def normalize_unit(value: Any) -> str:
    text = str(value or "").strip()
    return PLANT_ALIASES.get(text, text)


def slot_sort_key(slot: str) -> tuple[int, int]:
    match = re.match(r"(\d{1,2}):(\d{2})-", str(slot))
    if not match:
        return (99, 99)
    return (int(match.group(1)), int(match.group(2)))


def quarter_labels(slot: str) -> list[str]:
    match = re.match(r"(\d{1,2}):(\d{2})-(\d{1,2}|24):(\d{2})", str(slot))
    if not match:
        return []
    start_hour = int(match.group(1))
    labels = []
    for minute_offset in (15, 30, 45, 60):
        total_minutes = start_hour * 60 + minute_offset
        total_minutes %= 24 * 60
        labels.append(f"{total_minutes // 60:02d}:{total_minutes % 60:02d}")
    return labels


def read_workbook_bytes(content: bytes):
    return load_workbook(BytesIO(content), data_only=True)


def read_template(content: bytes):
    return load_workbook(BytesIO(content), data_only=False)


def header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column
    return headers


def choose_sheet(wb, spec: SourceSpec):
    required = {
        spec.unit_col,
        spec.date_col,
        spec.slot_col,
        spec.quantity_col,
        spec.price_col,
    }
    if spec.green_col:
        required.add(spec.green_col)

    if spec.preferred_sheet and spec.preferred_sheet in wb.sheetnames:
        ws = wb[spec.preferred_sheet]
        headers = header_map(ws)
        if required.issubset(headers):
            return ws, headers

    for ws in wb.worksheets:
        headers = header_map(ws)
        if required.issubset(headers):
            return ws, headers

    missing = "、".join(sorted(required))
    raise AppError(f"{spec.title} 未找到包含这些列的 sheet：{missing}")


def collect_source(content: bytes, spec: SourceSpec, target_date: str, report_sheet: str) -> dict[str, Any]:
    wb = read_workbook_bytes(content)
    ws, headers = choose_sheet(wb, spec)

    units = SHEET_UNITS[report_sheet]
    grouped: dict[str, dict[str, dict[str, Any]]] = {
        unit: {} for unit in units
    }
    seen_dates: set[str] = set()
    matched_rows = 0
    ignored_units: set[str] = set()
    green_blank_rows = 0

    for row in range(2, ws.max_row + 1):
        row_date = normalize_date(ws.cell(row, headers[spec.date_col]).value)
        if row_date:
            seen_dates.add(row_date)
        if row_date != target_date:
            continue

        unit = normalize_unit(ws.cell(row, headers[spec.unit_col]).value)
        if unit not in units:
            if unit:
                ignored_units.add(unit)
            continue

        slot = str(ws.cell(row, headers[spec.slot_col]).value or "").strip()
        if not slot:
            continue

        quantity = to_number(ws.cell(row, headers[spec.quantity_col]).value)
        price = to_number(ws.cell(row, headers[spec.price_col]).value)
        green = 0.0
        if spec.green_col:
            green_cell = ws.cell(row, headers[spec.green_col]).value
            if green_cell in (None, ""):
                green_blank_rows += 1
            green = to_number(green_cell)

        fee = quantity * (price + green)
        bucket = grouped[unit].setdefault(
            slot, {"quantity": 0.0, "fee": 0.0, "rows": 0}
        )
        bucket["quantity"] += quantity
        bucket["fee"] += fee
        bucket["rows"] += 1
        matched_rows += 1

    for unit, slots in grouped.items():
        for slot, values in slots.items():
            quantity = values["quantity"]
            if report_sheet == "光伏" and abs(quantity) < NEAR_ZERO_QUANTITY:
                values["quantity"] = 0.0
                values["fee"] = 0.0
                values["price"] = 0.0
                values["quarter_quantity"] = 0.0
                if spec.market_key == "export" and values.get("rows"):
                    values["keep_template_price"] = True
            else:
                values["price"] = values["fee"] / quantity if quantity else None
                values["quarter_quantity"] = quantity / 4 if quantity else 0

    date_sample = sorted(seen_dates)[:6]
    if len(seen_dates) > 6:
        date_sample.append("...")

    return {
        "marketKey": spec.market_key,
        "marketLabel": MARKET_LABELS[spec.market_key],
        "sourceTitle": spec.title,
        "reportSheet": report_sheet,
        "sheetName": ws.title,
        "grouped": grouped,
        "stats": {
            "matchedRows": matched_rows,
            "datesSeen": date_sample,
            "dateCount": len(seen_dates),
            "ignoredUnits": sorted(ignored_units)[:8],
            "greenBlankRows": green_blank_rows,
        },
    }


def find_unit_blocks(ws, units: tuple[str, ...]) -> dict[str, dict[str, Any]]:
    blocks: dict[str, dict[str, Any]] = {}
    all_starts = []
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(2, col).value or "").strip()
        next_value = str(ws.cell(2, col + 1).value or "").strip()
        if value and next_value == "省内优先":
            all_starts.append((value, col))

    for index, (unit, start_col) in enumerate(all_starts):
        if unit not in units:
            continue
        end_col = all_starts[index + 1][1] if index + 1 < len(all_starts) else ws.max_column + 1
        markets: dict[str, tuple[int, int]] = {}
        for col in range(start_col + 1, end_col):
            label = str(ws.cell(2, col).value or "").strip()
            sub_label = str(ws.cell(3, col).value or "").strip()
            if label in MARKET_LABELS.values() and sub_label == "电量":
                markets[label] = (col, col + 1)

        row_by_time: dict[str, int] = {}
        for row in range(4, ws.max_row + 1):
            label = normalize_time(ws.cell(row, start_col).value)
            if re.fullmatch(r"\d{2}:\d{2}", label):
                row_by_time[label] = row

        blocks[unit] = {
            "startCol": start_col,
            "endCol": end_col - 1,
            "markets": markets,
            "rowByTime": row_by_time,
        }

    return blocks


def reset_pv_market(
    ws,
    blocks: dict[str, dict[str, Any]],
    units: tuple[str, ...],
    market_label: str,
    value: float | None,
) -> int:
    reset_count = 0
    for unit in units:
        block = blocks.get(unit)
        if not block:
            continue
        rows = sorted(set(block["rowByTime"].values()))
        if market_label not in block["markets"]:
            continue
        quantity_col, price_col = block["markets"][market_label]
        for row in rows:
            quantity_cell = ws.cell(row, quantity_col)
            price_cell = ws.cell(row, price_col)
            quantity_cell.value = value
            quantity_cell.number_format = "0.00"
            price_cell.value = value
            price_cell.number_format = "0.00"
            reset_count += 2
    return reset_count


def snapshot_market_prices(
    wb,
    sheet_blocks: dict[str, dict[str, dict[str, Any]]],
) -> dict[tuple[str, str, str, int], Any]:
    prices: dict[tuple[str, str, str, int], Any] = {}
    for sheet_name, blocks in sheet_blocks.items():
        ws = wb[sheet_name]
        for unit, block in blocks.items():
            rows = sorted(set(block["rowByTime"].values()))
            for market_label, (_quantity_col, price_col) in block["markets"].items():
                for row in rows:
                    prices[(sheet_name, unit, market_label, row)] = ws.cell(row, price_col).value
    return prices


def find_summary_row(ws, block: dict[str, Any]) -> int:
    data_rows = sorted(set(block["rowByTime"].values()))
    last_data_row = max(data_rows)
    for row in range(last_data_row + 1, min(ws.max_row, last_data_row + 6) + 1):
        label = str(ws.cell(row, block["startCol"]).value or "").strip()
        if label in SUMMARY_LABELS:
            return row
    return last_data_row + 1


def restore_unwritten_template_prices(
    wb,
    sheet_blocks: dict[str, dict[str, dict[str, Any]]],
    template_prices: dict[tuple[str, str, str, int], Any],
    targets: set[tuple[str, str, str]],
    written_price_cells: set[tuple[str, str, str, int]],
) -> int:
    restored = 0
    for sheet_name, unit, market_label in targets:
        if sheet_name != "光伏":
            continue
        blocks = sheet_blocks.get(sheet_name, {})
        block = blocks.get(unit)
        if not block or market_label not in block["markets"]:
            continue
        ws = wb[sheet_name]
        quantity_col, price_col = block["markets"][market_label]
        for row in sorted(set(block["rowByTime"].values())):
            key = (sheet_name, unit, market_label, row)
            if key in written_price_cells:
                continue
            template_price = template_prices.get(key)
            if is_blank(template_price):
                continue
            quantity = to_number(ws.cell(row, quantity_col).value)
            if abs(quantity) >= NEAR_ZERO_QUANTITY:
                continue
            price_cell = ws.cell(row, price_col)
            price_cell.value = template_price
            price_cell.number_format = "0.00"
            restored += 1
    return restored


def apply_market_summaries(
    wb,
    sheet_blocks: dict[str, dict[str, dict[str, Any]]],
    targets: set[tuple[str, str, str]],
) -> int:
    updated = 0
    for sheet_name, unit, market_label in targets:
        blocks = sheet_blocks.get(sheet_name, {})
        block = blocks.get(unit)
        if not block or market_label not in block["markets"]:
            continue
        rows = sorted(set(block["rowByTime"].values()))
        if not rows:
            continue
        ws = wb[sheet_name]
        quantity_col, price_col = block["markets"][market_label]
        first_row = min(rows)
        last_row = max(rows)
        summary_row = find_summary_row(ws, block)
        quantity_letter = get_column_letter(quantity_col)
        price_letter = get_column_letter(price_col)

        quantity_cell = ws.cell(summary_row, quantity_col)
        price_cell = ws.cell(summary_row, price_col)
        quantity_cell.value = f"=SUM({quantity_letter}{first_row}:{quantity_letter}{last_row})"
        quantity_cell.number_format = "0.00"
        price_cell.value = f"={price_letter}{last_row}"
        price_cell.number_format = "0.00"
        updated += 2
    return updated


def apply_to_template(
    template_content: bytes,
    sources: list[dict[str, Any]],
    target_date: str,
    no_data_flags: set[str] | None = None,
) -> dict[str, Any]:
    no_data_flags = no_data_flags or set()
    wb = read_template(template_content)
    sheet_blocks: dict[str, dict[str, dict[str, Any]]] = {}
    protected_by_sheet: dict[str, dict[str, tuple[Any, str, Any]]] = {}

    for sheet_name in sorted({source["reportSheet"] for source in sources}):
        if sheet_name not in wb.sheetnames:
            raise AppError(f"上报模板未找到名为“{sheet_name}”的 sheet")
        ws = wb[sheet_name]
        protected_by_sheet[sheet_name] = snapshot_ranges(
            ws, PROTECTED_RANGES_BY_SHEET.get(sheet_name, ())
        )
        sheet_blocks[sheet_name] = find_unit_blocks(ws, SHEET_UNITS[sheet_name])

    fill_count = 0
    warnings: list[str] = []
    filled_cells: list[dict[str, Any]] = []
    template_prices = snapshot_market_prices(wb, sheet_blocks)
    active_market_targets: set[tuple[str, str, str]] = set()
    written_price_cells: set[tuple[str, str, str, int]] = set()

    if "光伏" in sheet_blocks:
        uploaded_keys = {source.get("inputKey") for source in sources}
        for input_key in no_data_flags:
            if input_key in uploaded_keys or not input_key.startswith("pv_"):
                continue
            config = SOURCE_INPUTS.get(input_key)
            if not config:
                continue
            market_label = MARKET_LABELS[SOURCE_SPECS[config["spec"]].market_key]
            fill_count += reset_pv_market(
                wb["光伏"],
                sheet_blocks["光伏"],
                INPUT_UNITS.get(input_key, SHEET_UNITS["光伏"]),
                market_label,
                None,
            )
        for source in sources:
            input_key = source.get("inputKey")
            if source["reportSheet"] != "光伏" or not input_key:
                continue
            units = INPUT_UNITS.get(input_key, SHEET_UNITS["光伏"])
            for unit in units:
                active_market_targets.add(("光伏", unit, source["marketLabel"]))
            fill_count += reset_pv_market(
                wb["光伏"],
                sheet_blocks["光伏"],
                units,
                source["marketLabel"],
                0,
            )

    for source in sources:
        report_sheet = source["reportSheet"]
        ws = wb[report_sheet]
        blocks = sheet_blocks[report_sheet]
        market_label = source["marketLabel"]
        grouped = source["grouped"]
        units_to_fill = INPUT_UNITS.get(source.get("inputKey"), SHEET_UNITS[report_sheet])
        for unit in units_to_fill:
            active_market_targets.add((report_sheet, unit, market_label))
            if unit not in blocks:
                warnings.append(f"{report_sheet} sheet 未找到 {unit} 列块")
                continue
            block = blocks[unit]
            if market_label not in block["markets"]:
                warnings.append(f"{report_sheet} {unit} 列块未找到 {market_label}")
                continue

            quantity_col, price_col = block["markets"][market_label]
            for slot, values in grouped[unit].items():
                labels = quarter_labels(slot)
                if not labels:
                    warnings.append(f"{source['sourceTitle']} {unit} 无法识别时段：{slot}")
                    continue

                missing_labels = [label for label in labels if label not in block["rowByTime"]]
                if missing_labels:
                    warnings.append(f"{report_sheet} {unit} {slot} 在模板中缺少 15 分钟点：{', '.join(missing_labels)}")
                    continue

                written_price = None
                for label in labels:
                    row = block["rowByTime"][label]
                    quantity_cell = ws.cell(row, quantity_col)
                    price_cell = ws.cell(row, price_col)
                    quantity_cell.value = round(values["quarter_quantity"], 2)
                    quantity_cell.number_format = "0.00"
                    if values.get("keep_template_price"):
                        price_cell.value = template_prices.get(
                            (report_sheet, unit, market_label, row),
                            price_cell.value,
                        )
                    else:
                        price_cell.value = None if values["price"] is None else round(values["price"], 2)
                    price_cell.number_format = "0.00"
                    written_price_cells.add((report_sheet, unit, market_label, row))
                    if written_price is None:
                        written_price = price_cell.value
                    fill_count += 2

                filled_cells.append(
                    {
                        "sheet": report_sheet,
                        "unit": unit,
                        "market": market_label,
                        "slot": slot,
                        "quantityCol": get_column_letter(quantity_col),
                        "priceCol": get_column_letter(price_col),
                        "quantity": values["quantity"],
                        "price": written_price,
                        "rows": values["rows"],
                    }
                )

    fill_count += restore_unwritten_template_prices(
        wb,
        sheet_blocks,
        template_prices,
        active_market_targets,
        written_price_cells,
    )
    fill_count += apply_market_summaries(wb, sheet_blocks, active_market_targets)

    for sheet_name, protected_cells in protected_by_sheet.items():
        restore_ranges(wb[sheet_name], protected_cells)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_date = target_date.replace("-", "")
    output_name = f"现货市场出清情况-金川_自动填报_{safe_date}_{timestamp}.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / output_name
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output_path)

    return {
        "outputName": output_name,
        "outputPath": str(output_path),
        "downloadUrl": f"/download/{output_name}",
        "fillCount": fill_count,
        "warnings": warnings,
        "filledCells": filled_cells,
    }


def snapshot_ranges(ws, ranges: tuple[str, ...]) -> dict[str, tuple[Any, str, Any]]:
    cells: dict[str, tuple[Any, str, Any]] = {}
    for range_text in ranges:
        for row in ws[range_text]:
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cells[cell.coordinate] = (cell.value, cell.number_format, copy.copy(cell._style))
    return cells


def restore_ranges(ws, cells: dict[str, tuple[Any, str, Any]]) -> None:
    for coordinate, (value, number_format, style) in cells.items():
        cell = ws[coordinate]
        cell.value = value
        cell.number_format = number_format
        cell._style = style


def build_preview(sources: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_unit_slot: dict[tuple[str, str, str], dict[str, Any]] = {}
    for source in sources:
        report_sheet = source["reportSheet"]
        market_label = source["marketLabel"]
        for unit, slots in source["grouped"].items():
            for slot, values in slots.items():
                row = by_unit_slot.setdefault(
                    (report_sheet, unit, slot),
                    {"sheet": report_sheet, "unit": unit, "slot": slot, "markets": {}},
                )
                row["markets"][market_label] = {
                    "quantity": round(values["quantity"], 2),
                    "quarterQuantity": round(values["quarter_quantity"], 2),
                    "price": None if values["price"] is None else round(values["price"], 2),
                    "rows": values["rows"],
                }

    return [
        by_unit_slot[key]
        for key in sorted(by_unit_slot, key=lambda item: (item[0], item[1], slot_sort_key(item[2])))
    ]


def process_files(files: dict[str, bytes], target_date: str, no_data_flags: set[str] | None = None) -> dict[str, Any]:
    no_data_flags = no_data_flags or set()
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", target_date or ""):
        raise AppError("请填写交易标的日期，格式为 YYYY-MM-DD")

    missing = [label for key, label in REQUIRED_FILES.items() if key not in files]
    if missing:
        raise AppError("缺少文件：" + "、".join(missing))

    sources = []
    for input_key, config in SOURCE_INPUTS.items():
        content = files.get(input_key)
        if not content:
            continue
        spec = SOURCE_SPECS[config["spec"]]
        source = collect_source(content, spec, target_date, config["sheet"])
        source["inputKey"] = input_key
        sources.append(source)

    if not sources and not no_data_flags:
        raise AppError("请至少上传一个待计算的数据源文件")

    output = apply_to_template(files["template"], sources, target_date, no_data_flags)

    checks = []
    for input_key, label in OPTIONAL_INPUT_LABELS.items():
        has_file = input_key in files
        marked_no_data = input_key in no_data_flags
        if has_file and marked_no_data:
            checks.append(
                {
                    "status": "warn",
                    "title": label,
                    "detail": "已上传文件，同时勾选了今日无数据；本次按上传文件计算",
                }
            )
        elif marked_no_data:
            detail = "已确认今日无数据，模板对应列保持原样"
            if input_key.startswith("pv_"):
                detail = "已确认今日无数据，光伏对应列已清空"
            checks.append(
                {
                    "status": "ok",
                    "title": label,
                    "detail": detail,
                }
            )
        elif not has_file:
            checks.append(
                {
                    "status": "warn",
                    "title": label,
                    "detail": "未上传，也未标记今日无数据",
                }
            )

    for source in sources:
        stats = source["stats"]
        status = "ok" if stats["matchedRows"] else "error"
        checks.append(
            {
                "status": status,
                "title": f"{source['reportSheet']} - {source['sourceTitle']}",
                "detail": f"{source['sheetName']} 匹配 {stats['matchedRows']} 行，识别日期 {stats['dateCount']} 个",
                "datesSeen": stats["datesSeen"],
            }
        )
        if stats["greenBlankRows"]:
            checks.append(
                {
                    "status": "warn",
                    "title": f"{source['sourceTitle']} 绿证价格",
                    "detail": f"{stats['greenBlankRows']} 行为空，已按 0 计算",
                }
            )

    checks.append(
        {
            "status": "ok" if output["fillCount"] else "error",
            "title": "模板回填",
            "detail": f"已写入 {output['fillCount']} 个单元格",
        }
    )
    for warning in output["warnings"][:12]:
        checks.append({"status": "warn", "title": "模板提示", "detail": warning})

    return {
        "targetDate": target_date,
        "checks": checks,
        "preview": build_preview(sources),
        "sources": [
            {
                "title": source["sourceTitle"],
                "sheet": source["reportSheet"],
                "market": source["marketLabel"],
                "stats": source["stats"],
            }
            for source in sources
        ],
        "output": output,
    }


def parse_json_body(handler: BaseHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length", "0") or 0)
    if not length:
        return {}
    raw = handler.rfile.read(length)
    if not raw:
        return {}
    return json.loads(raw.decode("utf-8"))


def parse_multipart(handler: BaseHTTPRequestHandler) -> tuple[dict[str, str], dict[str, bytes]]:
    content_type = handler.headers.get("Content-Type", "")
    if not content_type.lower().startswith("multipart/form-data"):
        raise AppError("请求格式错误，请使用网页上传文件")
    length = int(handler.headers.get("Content-Length", "0") or 0)
    raw_body = handler.rfile.read(length)
    message = BytesParser(policy=email_policy).parsebytes(
        b"Content-Type: " + content_type.encode("utf-8") + b"\r\n"
        b"MIME-Version: 1.0\r\n\r\n" + raw_body
    )
    if not message.is_multipart():
        raise AppError("请求格式错误，未读取到上传内容")

    fields: dict[str, str] = {}
    files: dict[str, bytes] = {}
    run_dir = UPLOAD_DIR / datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir.mkdir(parents=True, exist_ok=True)

    for part in message.iter_parts():
        if part.get_content_disposition() != "form-data":
            continue
        key = part.get_param("name", header="content-disposition")
        if not key:
            continue
        filename = part.get_filename()
        payload = part.get_payload(decode=True) or b""
        if filename:
            files[key] = payload
            safe_name = Path(filename).name
            (run_dir / safe_name).write_bytes(payload)
        else:
            charset = part.get_content_charset() or "utf-8"
            fields[key] = payload.decode(charset, errors="replace")

    return fields, files


def sample_file_bytes() -> dict[str, bytes]:
    missing = [str(path) for path in SAMPLE_FILES.values() if not path.exists()]
    if missing:
        raise AppError("示例文件不存在：" + "；".join(missing))
    return {key: path.read_bytes() for key, path in SAMPLE_FILES.items()}


def sample_87_file_bytes() -> dict[str, bytes]:
    missing = [str(path) for path in SAMPLE_87_FILES.values() if not path.exists()]
    if missing:
        raise AppError("8月7日示例文件不存在：" + "；".join(missing))
    return {key: path.read_bytes() for key, path in SAMPLE_87_FILES.items()}


def no_data_flags_from_fields(fields: dict[str, str]) -> set[str]:
    flags = set()
    for key, value in fields.items():
        if not key.startswith("no_data_"):
            continue
        if str(value).lower() in {"1", "true", "on", "yes"}:
            flags.add(key.removeprefix("no_data_"))
    return flags


def start_process_job(files: dict[str, bytes], target_date: str, no_data_flags: set[str]) -> str:
    job_id = uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[job_id] = {
            "id": job_id,
            "status": "running",
            "message": "文件已上传，正在计算",
            "createdAt": datetime.now().isoformat(timespec="seconds"),
        }

    def run_job() -> None:
        try:
            print(f"[{datetime.now():%H:%M:%S}] 后台任务开始：{job_id}", flush=True)
            result = process_files(files, target_date, no_data_flags)
            with JOBS_LOCK:
                JOBS[job_id].update(
                    {
                        "status": "done",
                        "message": "计算完成",
                        "finishedAt": datetime.now().isoformat(timespec="seconds"),
                        "result": result,
                    }
                )
            print(f"[{datetime.now():%H:%M:%S}] 后台任务完成：{job_id}", flush=True)
        except Exception as exc:
            detail = str(exc) if isinstance(exc, AppError) else "处理失败，请检查文件格式"
            print(f"[{datetime.now():%H:%M:%S}] 后台任务失败：{job_id}，{repr(exc)}", flush=True)
            traceback.print_exc()
            with JOBS_LOCK:
                JOBS[job_id].update(
                    {
                        "status": "error",
                        "message": "计算失败",
                        "finishedAt": datetime.now().isoformat(timespec="seconds"),
                        "error": detail,
                    }
                )

    threading.Thread(target=run_job, name=f"process-{job_id[:8]}", daemon=True).start()
    return job_id


class MarketReportHandler(BaseHTTPRequestHandler):
    server_version = "MarketReportWebApp/1.0"

    def do_GET(self):
        path = unquote(urlparse(self.path).path)
        try:
            if path == "/":
                self.send_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
            elif path == "/api/health":
                self.send_json({"ok": True, "python": sys.executable})
            elif path.startswith("/static/"):
                file_path = STATIC_DIR / path.removeprefix("/static/")
                self.send_file(file_path)
            elif path.startswith("/api/jobs/"):
                job_id = Path(path.removeprefix("/api/jobs/")).name
                with JOBS_LOCK:
                    job = JOBS.get(job_id)
                    payload = copy.deepcopy(job) if job else None
                if not payload:
                    raise AppError("任务不存在或已过期，请重新上传计算")
                self.send_json(payload)
            elif path.startswith("/download/"):
                file_name = Path(path.removeprefix("/download/")).name
                self.send_file(OUTPUT_DIR / file_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_name=file_name)
            else:
                self.send_error(404, "Not found")
        except Exception as exc:
            self.send_error_json(exc)

    def do_POST(self):
        path = unquote(urlparse(self.path).path)
        try:
            if path == "/api/process":
                fields, files = parse_multipart(self)
                job_id = start_process_job(
                    files,
                    fields.get("target_date", ""),
                    no_data_flags_from_fields(fields),
                )
                self.send_json({"jobId": job_id, "status": "running", "message": "文件已上传，正在后台计算"}, 202)
            elif path == "/api/process-sample":
                payload = parse_json_body(self)
                result = process_files(sample_file_bytes(), payload.get("target_date", ""))
                self.send_json(result)
            elif path == "/api/process-sample-87":
                payload = parse_json_body(self)
                files = sample_87_file_bytes()
                no_data = set(OPTIONAL_INPUT_LABELS) - set(files)
                result = process_files(files, payload.get("target_date", ""), no_data)
                self.send_json(result)
            else:
                self.send_error(404, "Not found")
        except Exception as exc:
            self.send_error_json(exc)

    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"[{datetime.now():%H:%M:%S}] {self.address_string()} {fmt % args}")

    def send_json(self, data: dict[str, Any], status: int = 200):
        body = json.dumps(data, ensure_ascii=False, default=str).encode("utf-8")
        try:
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError, OSError) as exc:
            print(f"[{datetime.now():%H:%M:%S}] 客户端连接已断开，响应未发送：{exc}")

    def send_error_json(self, exc: Exception):
        detail = str(exc) if isinstance(exc, AppError) else "处理失败，请检查文件格式"
        print(f"[{datetime.now():%H:%M:%S}] 处理请求失败：{repr(exc)}")
        traceback.print_exc()
        self.send_json({"error": detail}, 400)

    def send_file(self, path: Path, content_type: str | None = None, download_name: str | None = None):
        if not path.exists() or not path.is_file():
            raise AppError("文件不存在")
        if content_type is None:
            suffix = path.suffix.lower()
            content_type = {
                ".html": "text/html; charset=utf-8",
                ".css": "text/css; charset=utf-8",
                ".js": "application/javascript; charset=utf-8",
                ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }.get(suffix, "application/octet-stream")

        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(path.stat().st_size))
        self.send_header("Cache-Control", "no-store")
        if download_name:
            encoded_name = quote(download_name)
            self.send_header(
                "Content-Disposition",
                f"attachment; filename=report.xlsx; filename*=UTF-8''{encoded_name}",
            )
        self.end_headers()
        with path.open("rb") as handle:
            shutil.copyfileobj(handle, self.wfile)


def main():
    cloud_mode = "PORT" in os.environ
    host = os.environ.get("HOST", "0.0.0.0" if cloud_mode else "127.0.0.1")
    port = int(os.environ.get("PORT", "8765"))
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    httpd = ThreadingHTTPServer((host, port), MarketReportHandler)
    local_url = f"http://127.0.0.1:{port}" if host == "0.0.0.0" else f"http://{host}:{port}"
    print(f"电力市场报表自动填报工具已启动：{local_url}", flush=True)
    print("按 Ctrl+C 停止服务", flush=True)
    httpd.serve_forever()


if __name__ == "__main__":
    main()
